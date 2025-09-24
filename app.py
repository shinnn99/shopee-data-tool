import gradio as gr
import pandas as pd
import os
import time
import re
import json

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Color
from openpyxl.utils import get_column_letter

# Thư viện cho Google Sheets
import gspread
from gspread.exceptions import WorksheetNotFound
from gspread_dataframe import set_with_dataframe
from google.oauth2.service_account import Credentials

# --- HÀM KIỂM TRA SHEET TRỐNG THÔNG MINH ---
def is_sheet_truly_empty(data):
    """Kiểm tra xem sheet có thực sự trống không (kể cả các hàng trống)."""
    if not data:
        return True
    return all(all(cell == '' for cell in row) for row in data)

# --- HÀM TẢI LÊN GOOGLE SHEET VỚI LOGIC MỚI ---
def upload_to_google_sheet(df, sheet_url, sheet_name, progress):
    try:
        progress(0.9, desc="Đang kết nối tới Google Sheets...")
        scopes = ['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive']
        creds_json_str = os.getenv("GOOGLE_CREDENTIALS_JSON")
        if not creds_json_str:
            return "❌ Lỗi: Không tìm thấy biến môi trường 'GOOGLE_CREDENTIALS_JSON' trên server."
        creds_info = json.loads(creds_json_str)
        creds = Credentials.from_service_account_info(creds_info, scopes=scopes)
        client = gspread.authorize(creds)
        spreadsheet = client.open_by_url(sheet_url)
        
        target_worksheet = None
        final_sheet_name = ""

        # === LOGIC MỚI BẮT ĐẦU ===
        if sheet_name:  # Trường hợp 1: Người dùng nhập tên sheet cụ thể
            final_sheet_name = sheet_name
            try:
                target_worksheet = spreadsheet.worksheet(final_sheet_name)
                progress(0.92, desc=f"Đã tìm thấy sheet '{final_sheet_name}'.")
            except WorksheetNotFound:
                progress(0.92, desc=f"Không tìm thấy sheet '{final_sheet_name}', đang tạo mới...")
                target_worksheet = spreadsheet.add_worksheet(title=final_sheet_name, rows="1", cols="20")
        else:  # Trường hợp 2: Người dùng không nhập tên sheet -> áp dụng logic tự động
            try:
                sheet1 = spreadsheet.worksheet("Sheet1")
                sheet1_data = sheet1.get_all_values()
                if is_sheet_truly_empty(sheet1_data):
                    target_worksheet = sheet1
                    final_sheet_name = "Sheet1"
                    progress(0.92, desc="Sheet1 trống, sẽ ghi dữ liệu vào đây.")
                else:
                    progress(0.92, desc="Sheet1 đã có dữ liệu, đang tạo sheet mới 'file_gia'...")
                    # Tìm tên sheet mới không bị trùng
                    new_sheet_base_name = "file_gia"
                    new_sheet_name = new_sheet_base_name
                    i = 1
                    while True:
                        try:
                            spreadsheet.worksheet(new_sheet_name)
                            new_sheet_name = f"{new_sheet_base_name}_{i}"
                            i += 1
                        except WorksheetNotFound:
                            break
                    final_sheet_name = new_sheet_name
                    target_worksheet = spreadsheet.add_worksheet(title=final_sheet_name, rows="1", cols="20")
            except WorksheetNotFound:
                progress(0.92, desc="Không tìm thấy Sheet1, đang tạo mới...")
                final_sheet_name = "Sheet1"
                target_worksheet = spreadsheet.add_worksheet(title=final_sheet_name, rows="1", cols="20")
        # === LOGIC MỚI KẾT THÚC ===

        progress(0.95, desc=f"Đang ghi dữ liệu vào sheet '{final_sheet_name}'...")
        time.sleep(1)
        
        existing_data = target_worksheet.get_all_values()
        rows_to_append = df.values.tolist()
        
        if is_sheet_truly_empty(existing_data):
            header = df.columns.tolist()
            target_worksheet.append_row(header, value_input_option='USER_ENTERED')
            target_worksheet.append_rows(rows_to_append, value_input_option='USER_ENTERED')
            header_format = {"textFormat": {"bold": True}}
            target_worksheet.format('A1:{}'.format(get_column_letter(len(df.columns)) + '1'), header_format)
            return f"✅ Thành công! Đã tạo và ghi dữ liệu vào sheet '{final_sheet_name}'."
        else: # Trường hợp này chỉ xảy ra khi người dùng tự nhập tên sheet đã có dữ liệu
            target_worksheet.append_rows(rows_to_append, value_input_option='USER_ENTERED')
            return f"✅ Thành công! Đã nối thêm {len(rows_to_append)} dòng mới vào sheet '{final_sheet_name}'."

    except gspread.exceptions.SpreadsheetNotFound:
        return "❌ Lỗi: Không tìm thấy Google Sheet. Vui lòng kiểm tra lại đường link hoặc quyền chia sẻ."
    except Exception as e:
        return f"❌ Lỗi khi tải lên Google Sheets: {e}"

# --- HÀM XỬ LÝ DỮ LIỆU CHÍNH (KHÔNG THAY ĐỔI) ---
def process_data(shop_id_input, source_files, font_name, font_size, output_choice, gsheet_url, sheet_name, progress=gr.Progress()):
    progress(0, desc="Đang kiểm tra thông tin...")
    if not shop_id_input: return "❌ Lỗi: Vui lòng nhập Shop ID!", None
    if not source_files: return "❌ Lỗi: Vui lòng tải lên ít nhất một file dữ liệu nguồn!", None
    if output_choice == "Tải lên Google Sheet" and not gsheet_url: return "❌ Lỗi: Vui lòng nhập đường link Google Sheet!", None
    progress(0.1, desc="Đang xác thực Shop ID...")
    shop_id = str(shop_id_input).strip()
    if not shop_id.isdigit(): return f"❌ Lỗi: Shop ID '{shop_id}' không hợp lệ. Vui lòng chỉ nhập số.", None
    all_data_frames = []
    total_files = len(source_files)
    for i, source_file in enumerate(source_files):
        progress(0.2 + (i / total_files) * 0.3, desc=f"Đang đọc file {i+1}/{total_files}...")
        try:
            file_path = source_file.name
            file_extension = os.path.splitext(file_path)[1].lower()
            if file_extension in ('.xlsx', '.xls'): df_source = pd.read_excel(file_path)
            elif file_extension == '.csv': df_source = pd.read_csv(file_path)
            else: return f"❌ Lỗi: File {os.path.basename(file_path)} có định dạng không được hỗ trợ.", None
            df_output = pd.DataFrame()
            column_mapping = { 'ID sản phẩm': 'Mã sản phẩm', 'Tên sản phẩm': 'Tên Sản phẩm (Tùy chọn)', 'ID phân loại': 'Mã phân loại hàng', 'Tên phân loại': 'Tên phân loại hàng (Tùy chọn)', 'Giá gốc': 'Giá gốc (Tùy chọn)', 'Giá đang bán': 'Giá đã giảm' }
            price_cols_source = ['Giá gốc (Tùy chọn)', 'Giá đã giảm']
            for col in price_cols_source:
                if col in df_source.columns: df_source[col] = pd.to_numeric(df_source[col], errors='coerce').fillna(0)
            id_cols_source = ['Mã sản phẩm', 'Mã phân loại hàng']
            for col in id_cols_source:
                if col in df_source.columns: df_source[col] = df_source[col].astype(str).str.replace(r'\.0$', '', regex=True)
            for template_col, source_col in column_mapping.items():
                if source_col in df_source.columns: df_output[template_col] = df_source[source_col]
                else: df_output[template_col] = ''
            df_output['Shop_id'] = shop_id
            df_output['Link'] = 'https://shopee.vn/a-i.' + df_output['Shop_id'].astype(str) + '.' + df_output['ID sản phẩm'].astype(str)
            df_output['Giá FS'] = ''
            df_output['Giá campaign'] = ''
            all_data_frames.append(df_output)
        except Exception as e: return f"❌ Lỗi khi đọc hoặc xử lý file {os.path.basename(source_file.name)}: {e}", None
    if not all_data_frames: return "❌ Lỗi: Không có dữ liệu nào được xử lý.", None
    progress(0.5, desc="Đang gộp dữ liệu từ các file...")
    final_df = pd.concat(all_data_frames, ignore_index=True)
    final_columns_order = [ 'Shop_id', 'ID sản phẩm', 'Tên sản phẩm', 'ID phân loại', 'Tên phân loại', 'Link', 'Giá gốc', 'Giá đang bán', 'Giá FS', 'Giá campaign' ]
    final_df = final_df.reindex(columns=final_columns_order, fill_value='')

    if output_choice == "Tải xuống Excel":
        progress(0.8, desc="Đang định dạng và lưu file Excel...")
        try:
            output_filepath = f"template_final_{shop_id}.xlsx"
            writer = pd.ExcelWriter(output_filepath, engine='openpyxl')
            final_df.to_excel(writer, index=False, sheet_name='Sheet1')
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']

            header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
            header_font = Font(name=font_name, size=font_size, bold=True, color="FFFFFF")
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            data_font = Font(name=font_name, size=font_size - 1)
            data_alignment = Alignment(vertical='center', wrap_text=False) 
            
            for col_idx, column_cell in enumerate(worksheet.columns, 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                
                max_length = 0
                column_letter = get_column_letter(col_idx)
                column_header = str(cell.value)

                for cell_in_col in column_cell:
                    if cell_in_col.row > 1:
                        cell_in_col.font = data_font
                        cell_in_col.alignment = data_alignment
                    try:
                        if len(str(cell_in_col.value)) > max_length:
                            max_length = len(str(cell_in_col.value))
                    except:
                        pass
                
                if column_header == "Tên sản phẩm":
                    adjusted_width = 45
                elif column_header == "Link":
                    adjusted_width = 30
                else:
                    adjusted_width = (max_length + 2) * 1.2
                
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            worksheet.freeze_panes = 'A2'
            writer.close()
            progress(1, desc="Hoàn thành!")
            return f"✅ Thành công! File Excel đã được tạo.", output_filepath
        except Exception as e: return f"❌ Lỗi khi lưu file Excel: {e}", None
            
    elif output_choice == "Tải lên Google Sheet":
        status_message = upload_to_google_sheet(final_df, gsheet_url, sheet_name, progress)
        progress(1, desc="Hoàn thành!")
        return status_message, None

# --- GIAO DIỆN GRADIO (KHÔNG THAY ĐỔI) ---
with gr.Blocks(theme=gr.themes.Soft(), title="Công cụ tạo file giá") as demo:
    gr.HTML("<h1 style='text-align: center; color: #107C41;'>Công cụ tạo file giá Shopee ✨</h1>")
    gr.Markdown("<p style='text-align: center;'>Nhập Shop ID và file dữ liệu nguồn để tạo file mẫu nhanh chóng.</p>")

    with gr.Row():
        with gr.Column(scale=2): # Cột trái cho Input
            with gr.Group():
                gr.Markdown("### 1. Thông tin bắt buộc")
                shop_id_input = gr.Textbox(label="Shop ID Shopee:", placeholder="Nhập chính xác ID của Shop (chỉ gồm số)...")
                file_input = gr.File(label="Tải lên file dữ liệu nguồn (.xlsx, .csv):", file_types=[".xlsx", ".xls", ".csv"], file_count="multiple")
            
            with gr.Group():
                gr.Markdown("### 2. Lựa chọn Output")
                output_choice_radio = gr.Radio(label="Lưu kết quả ở đâu?", choices=["Tải xuống Excel", "Tải lên Google Sheet"], value="Tải xuống Excel")
                
                with gr.Group(visible=False) as gsheet_group:
                    google_sheet_url_input = gr.Textbox(label="Đường link Google Sheet:", placeholder="Dán link Google Sheet của bạn vào đây...")
                    sheet_name_input = gr.Textbox(label="Tên Sheet mong muốn:", placeholder="Mặc định: Tự động", info="Để trống để tự động ghi vào Sheet1 (nếu trống) hoặc tạo sheet 'file_gia' (nếu Sheet1 đã có dữ liệu).")

            with gr.Group():
                gr.Markdown("### 3. Tùy chọn định dạng (chỉ cho Excel)")
                with gr.Row():
                    font_name_dropdown = gr.Dropdown(choices=["Calibri", "Arial", "Times New Roman"], value="Calibri", label="Font chữ")
                    font_size_slider = gr.Slider(minimum=8, maximum=24, step=1, value=12, label="Cỡ chữ")

        with gr.Column(scale=1): # Cột phải cho Output
            process_button = gr.Button("🚀 Bắt đầu Xử lý", variant="primary", scale=1)
            status_output = gr.Textbox(label="Trạng thái:", interactive=False, lines=3)
            file_output = gr.File(label="File Kết quả (Excel):")

    # --- Event Handlers ---
    def toggle_gsheet_url_visibility(choice):
        is_visible = (choice == "Tải lên Google Sheet")
        return gr.update(visible=is_visible)

    output_choice_radio.change(
        fn=toggle_gsheet_url_visibility,
        inputs=output_choice_radio,
        outputs=[gsheet_group]
    )
    
    process_button.click(
        fn=process_data,
        inputs=[shop_id_input, file_input, font_name_dropdown, font_size_slider, output_choice_radio, google_sheet_url_input, sheet_name_input],
        outputs=[status_output, file_output]
    )

# --- Khởi chạy app cho server ---
port = int(os.getenv('PORT', 7860))
demo.launch(server_name="0.0.0.0", server_port=port)
